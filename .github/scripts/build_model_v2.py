from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from collections import defaultdict, Counter
from pathlib import Path
from datetime import datetime, date
import argparse
import base64
import gzip
import hashlib
import json
import math
import re
import unicodedata

MONTHLY = [
    (15, 16, 17, 18, 19),
    (20, 21, 22, 23, 24),
    (25, 26, 27, 28, 29),
    (30, 31, 32, 33, 34),
    (35, 36, 37, 38, 39),
    (40, 41, 42, 43, 44),
]


def number(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        value = float(value)
        return value if math.isfinite(value) else 0.0
    raw = str(value).strip()
    try:
        if "," in raw:
            raw = raw.replace(".", "").replace(",", ".")
        value = float(raw)
        return value if math.isfinite(value) else 0.0
    except Exception:
        return 0.0


def clean_number(value):
    value = round(number(value), 10)
    if abs(value) < 1e-12:
        return 0
    if abs(value - round(value)) < 1e-10:
        return int(round(value))
    return value


def clean_text(value, default=""):
    value = "" if value is None else str(value).strip()
    return value or default


def normalize(value):
    value = "" if value is None else str(value)
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", value.lower().strip())


def parse_date(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    raw = clean_text(value)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            pass
    return None


def git_blob_sha(text):
    raw = text.encode("utf-8")
    return hashlib.sha1(b"blob " + str(len(raw)).encode("ascii") + b"\0" + raw).hexdigest()


def build(model_path, estudo_path, output_path):
    output_path.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(model_path, data_only=True, read_only=False)
    if "Planilha1" not in workbook.sheetnames:
        raise RuntimeError("A aba Planilha1 não foi localizada na Modelo Cronograma Base de Dados  v2.xlsx")
    sheet = workbook["Planilha1"]

    expected_headers = {
        2: "Codigo e Nome da Obra",
        3: "Item/CPU - SIENGE",
        4: "Unid serviço/Coef Insumo",
        5: "Preço Unit",
        10: "Quant. Total JULHO 2026",
    }
    for column, expected in expected_headers.items():
        actual = clean_text(sheet.cell(1, column).value)
        if actual != expected:
            raise RuntimeError(
                f"Cabeçalho divergente na coluna {column}: {actual!r} != {expected!r}"
            )

    expected_month_headers = [
        "Qnt Realis.",
        "R$ Realista",
        "Qtd. Executada",
        "R$ Executado",
        "Ordem de Serviço",
    ]
    for group in MONTHLY:
        actual = [clean_text(sheet.cell(2, column).value) for column in group]
        if actual != expected_month_headers:
            raise RuntimeError(f"Cabeçalhos mensais divergentes: {actual!r}")

    order_columns = {group[4] for group in MONTHLY}
    merged_orders = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_col == max_col and min_col in order_columns:
            value = sheet.cell(min_row, min_col).value
            for row in range(min_row, max_row + 1):
                merged_orders[(row, min_col)] = value

    contract_data = defaultdict(list)
    occurrences = defaultdict(Counter)
    code_to_contract = {}
    source_by_contract_item = defaultdict(list)
    unit_by_contract_source = {}

    for row in range(4, sheet.max_row + 1):
        code_raw = sheet.cell(row, 1).value
        contract = clean_text(sheet.cell(row, 2).value)
        item = clean_text(sheet.cell(row, 3).value)
        unit = clean_text(sheet.cell(row, 4).value, "-")
        price_raw = sheet.cell(row, 5).value

        if contract and code_raw not in (None, ""):
            try:
                code = str(int(float(code_raw)))
            except Exception:
                code = clean_text(code_raw)
            code_to_contract[code] = contract

        if (
            not contract
            or not item
            or not isinstance(price_raw, (int, float))
            or abs(float(price_raw)) < 1e-12
        ):
            continue

        occurrences[contract][item] += 1
        occurrence = occurrences[contract][item]
        source_id = item if occurrence == 1 else f"{item} #{occurrence}"

        planned = [0] * 12
        planned_value = [0] * 12
        executed = [0] * 12
        executed_value = [0] * 12
        orders = ["Sem serviços"] * 12

        for month, (
            plan_column,
            plan_value_column,
            exec_column,
            exec_value_column,
            order_column,
        ) in enumerate(MONTHLY):
            planned[month] = clean_number(sheet.cell(row, plan_column).value)
            planned_value[month] = clean_number(
                sheet.cell(row, plan_value_column).value
            )
            executed[month] = clean_number(sheet.cell(row, exec_column).value)
            executed_value[month] = clean_number(
                sheet.cell(row, exec_value_column).value
            )
            orders[month] = clean_text(
                merged_orders.get(
                    (row, order_column), sheet.cell(row, order_column).value
                ),
                "Sem serviços",
            )

        contract_data[contract].append(
            {
                "sourceId": source_id,
                "name": item,
                "unit": unit,
                "price": clean_number(price_raw),
                "contractual": clean_number(sheet.cell(row, 8).value),
                "base": clean_number(sheet.cell(row, 12).value),
                "balance": clean_number(sheet.cell(row, 10).value),
                "plan": planned,
                "planValue": planned_value,
                "exec": executed,
                "execValue": executed_value,
                "orders": orders,
            }
        )
        source_by_contract_item[(contract, normalize(item))].append(source_id)
        unit_by_contract_source[(contract, source_id)] = unit

    contract_data = dict(contract_data)
    contract_count = len(contract_data)
    item_count = sum(len(items) for items in contract_data.values())
    if contract_count != 15 or item_count != 650:
        raise RuntimeError(
            f"Importação divergente: {contract_count} contratos e {item_count} itens"
        )

    estudo = load_workbook(estudo_path, data_only=True, read_only=True)
    if "Base BI Executado (2)" not in estudo.sheetnames:
        raise RuntimeError(
            "A aba Base BI Executado (2) não foi localizada no Estudo Geral"
        )
    rdo_sheet = estudo["Base BI Executado (2)"]

    rdo_data = []
    unmatched_rdo = []
    for row_index, row in enumerate(
        rdo_sheet.iter_rows(min_row=4, values_only=True), start=4
    ):
        date_raw, code_raw, task_raw, _labor_price, _material_price, qty_raw = row[:6]
        if (
            date_raw in (None, "")
            or code_raw in (None, "")
            or task_raw in (None, "")
            or qty_raw in (None, "")
        ):
            continue

        date_value = parse_date(date_raw)
        quantity = number(qty_raw)
        if date_value is None or not math.isfinite(quantity):
            continue

        code = clean_text(code_raw)
        if code.endswith(".0"):
            code = code[:-2]
        contract_key = code_to_contract.get(code)
        task = clean_text(task_raw)

        if not contract_key:
            unmatched_rdo.append(
                {"row": row_index, "reason": "contract", "code": code, "task": task}
            )
            continue

        candidates = source_by_contract_item.get(
            (contract_key, normalize(task)), []
        )
        if not candidates:
            unmatched_rdo.append(
                {
                    "row": row_index,
                    "reason": "item",
                    "contract": contract_key,
                    "task": task,
                }
            )
            continue

        source_id = candidates[0]
        rdo_data.append(
            {
                "contract": contract_key.split(" - ")[0],
                "month": date_value.month - 1,
                "itemId": source_id,
                "observation": f"RDO de {date_value.strftime('%d/%m/%Y')}",
                "service": task,
                "quantity": clean_number(quantity),
                "unit": unit_by_contract_source.get((contract_key, source_id), "-"),
            }
        )

    if unmatched_rdo:
        raise RuntimeError(
            "Registros de RDO sem vínculo: "
            + json.dumps(unmatched_rdo, ensure_ascii=False)
        )
    if len(rdo_data) != 18:
        raise RuntimeError(
            f"Quantidade de registros RDO divergente: {len(rdo_data)}"
        )

    metadata = {
        "source": "Modelo Cronograma Base de Dados  v2.xlsx",
        "sheet": "Planilha1",
        "balanceColumn": "J - Quant. Total JULHO 2026",
        "unitPriceColumn": "E - Preço Unit",
        "historicalPlanQuantity": "Qnt Realis.",
        "historicalPlanValue": "R$ Realista",
        "historicalExecQuantity": "Qtd. Executada",
        "historicalExecValue": "R$ Executado",
        "futurePlanValueRule": "quantidade inserida no HTML x Preço Unit",
        "rdoSource": "Estudo Geral.xlsx / Base BI Executado (2)",
        "rdoRule": "somente quantidade; não altera o Executado do painel",
        "contracts": contract_count,
        "items": item_count,
        "rdoRecords": len(rdo_data),
    }

    script = (
        "var contractData="
        + json.dumps(contract_data, ensure_ascii=False, separators=(",", ":"))
        + ";\n"
    )
    script += (
        "var rdoData="
        + json.dumps(rdo_data, ensure_ascii=False, separators=(",", ":"))
        + ";\n"
    )
    script += (
        "var modelV2Metadata="
        + json.dumps(metadata, ensure_ascii=False, separators=(",", ":"))
        + ";\n"
    )

    packed = base64.b64encode(
        gzip.compress(script.encode("utf-8"), compresslevel=9, mtime=0)
    ).decode("ascii")
    part_size = (len(packed) + 4) // 5
    parts = []
    for index in range(5):
        part = packed[index * part_size : (index + 1) * part_size]
        (output_path / f"model-v2.part{index}").write_text(
            part, encoding="ascii"
        )
        parts.append(part)

    joined = "".join(parts)
    decoded = gzip.decompress(base64.b64decode(joined)).decode("utf-8")
    if decoded != script:
        raise RuntimeError("Falha na validação do pacote interno")

    monthly_totals = []
    for month in range(6):
        monthly_totals.append(
            {
                "month": month + 1,
                "plannedQuantity": round(
                    sum(
                        float(item["plan"][month] or 0)
                        for items in contract_data.values()
                        for item in items
                    ),
                    10,
                ),
                "plannedValue": round(
                    sum(
                        float(item["planValue"][month] or 0)
                        for items in contract_data.values()
                        for item in items
                    ),
                    10,
                ),
                "executedQuantity": round(
                    sum(
                        float(item["exec"][month] or 0)
                        for items in contract_data.values()
                        for item in items
                    ),
                    10,
                ),
                "executedValue": round(
                    sum(
                        float(item["execValue"][month] or 0)
                        for items in contract_data.values()
                        for item in items
                    ),
                    10,
                ),
            }
        )

    audit = {
        "metadata": metadata,
        "partLengths": [len(part) for part in parts],
        "partGitBlobShas": [git_blob_sha(part) for part in parts],
        "packedLength": len(packed),
        "scriptLength": len(script),
        "monthlyTotals": monthly_totals,
        "rdoByMonth": dict(Counter(record["month"] + 1 for record in rdo_data)),
        "rdoContracts": dict(Counter(record["contract"] for record in rdo_data)),
        "sampleRdo": rdo_data[:3],
    }
    (output_path / "model-v2-audit.json").write_text(
        json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(audit, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--model", required=True, type=Path)
    parser.add_argument("--estudo", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    arguments = parser.parse_args()
    build(arguments.model, arguments.estudo, arguments.output)
