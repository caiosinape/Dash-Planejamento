from openpyxl import load_workbook
from collections import defaultdict
from pathlib import Path
import json, math, re, unicodedata

repo = Path('.')
html_path = repo / 'Planejamento-Colaborativo/index.html'
source_path = repo / 'Cronogramas Planejamento/index.html'
data_path = repo / 'Planejamento-Colaborativo/model-v2-data.js'
html = html_path.read_text(encoding='utf-8')
source_html = source_path.read_text(encoding='utf-8')


def extract_json(text, variable):
    marker = 'var ' + variable + '='
    start = text.find(marker)
    if start < 0:
        return None
    start += len(marker)
    opening = text[start]
    closing = '}' if opening == '{' else ']'
    depth = 0
    quoted = False
    escaped = False
    for i in range(start, len(text)):
        ch = text[i]
        if quoted:
            if escaped:
                escaped = False
            elif ch == '\\':
                escaped = True
            elif ch == '"':
                quoted = False
        else:
            if ch == '"':
                quoted = True
            elif ch == opening:
                depth += 1
            elif ch == closing:
                depth -= 1
                if depth == 0:
                    return json.loads(text[start:i + 1])
    return None


def norm(value):
    value = unicodedata.normalize('NFKD', str(value or '')).encode('ascii', 'ignore').decode().lower()
    return re.sub(r'\s+', ' ', value).strip()


def number(value):
    try:
        value = float(value or 0)
        return value if math.isfinite(value) else 0.0
    except Exception:
        return 0.0


def text(value):
    return '' if value is None else str(value).strip()


old = extract_json(source_html, 'contractData') or {}
old_lookup = defaultdict(list)
for contract, rows in old.items():
    for row in rows:
        old_lookup[(contract, norm(row.get('name')))].append(row)

wb = load_workbook('/tmp/modelo-v2.xlsx', data_only=True, read_only=False)
ws = wb['Planilha1']

plan_qty_cols = [15, 20, 25, 30, 35, 40]
plan_value_cols = [16, 21, 26, 31, 36, 41]
exec_qty_cols = [17, 22, 27, 32, 37, 42]
exec_value_cols = [18, 23, 28, 33, 38, 43]
os_cols = [19, 24, 29, 34, 39, 44]

merged_os = {}
for merged in ws.merged_cells.ranges:
    if merged.min_col == merged.max_col and merged.min_col in os_cols:
        value = ws.cell(merged.min_row, merged.min_col).value
        for row_index in range(merged.min_row, merged.max_row + 1):
            merged_os[(row_index, merged.min_col)] = value


def monthly_os(row_index, col):
    return text(merged_os.get((row_index, col), ws.cell(row_index, col).value))


contracts = defaultdict(list)
occurrences = defaultdict(lambda: defaultdict(int))
skipped_headers = 0

for row_index in range(3, ws.max_row + 1):
    code = text(ws.cell(row_index, 1).value)
    contract = text(ws.cell(row_index, 2).value)
    name = text(ws.cell(row_index, 3).value)
    if not contract or not name:
        continue

    raw_unit = ws.cell(row_index, 4).value
    unit = text(raw_unit) if raw_unit not in (None, 0, '0') else ''
    price_unit = number(ws.cell(row_index, 5).value)
    balance_j = number(ws.cell(row_index, 10).value)
    accumulated_l = number(ws.cell(row_index, 12).value)

    plan = [0.0] * 12
    plan_value = [0.0] * 12
    executed = [0.0] * 12
    exec_value = [0.0] * 12
    orders = [''] * 12

    for month in range(6):
        plan[month] = number(ws.cell(row_index, plan_qty_cols[month]).value)
        plan_value[month] = number(ws.cell(row_index, plan_value_cols[month]).value)
        executed[month] = number(ws.cell(row_index, exec_qty_cols[month]).value)
        exec_value[month] = number(ws.cell(row_index, exec_value_cols[month]).value)
        orders[month] = monthly_os(row_index, os_cols[month])

    has_activity = any(abs(value) > 1e-12 for value in plan + plan_value + executed + exec_value)
    if not (unit or price_unit or balance_j or accumulated_l or has_activity):
        skipped_headers += 1
        continue

    occurrences[contract][name] += 1
    occurrence = occurrences[contract][name]
    fallback_list = old_lookup.get((contract, norm(name)), [])
    fallback = fallback_list[min(occurrence - 1, len(fallback_list) - 1)] if fallback_list else {}
    source_id = str(fallback.get('sourceId') or f'v2-r{row_index}')

    contracts[contract].append({
        'sourceId': source_id,
        'sourceRow': row_index,
        'code': code,
        'name': name,
        'unit': unit or text(fallback.get('unit')) or 'un.',
        'price': price_unit,
        'contractual': accumulated_l + balance_j,
        'base': accumulated_l,
        'balance': balance_j,
        'plan': plan,
        'planValue': plan_value,
        'exec': executed,
        'execValue': exec_value,
        'orders': orders,
        'occurrence': occurrence,
    })

model = {
    'metadata': {
        'source': 'Modelo Cronograma Base de Dados v2.xlsx',
        'sheet': 'Planilha1',
        'updatedAt': '2026-07-29T18:00:36Z',
        'contracts': len(contracts),
        'items': sum(len(rows) for rows in contracts.values()),
        'skippedCategoryRows': skipped_headers,
        'months': 'Janeiro a Junho/2026',
        'mapping': {
            'contract': 'B - Codigo e Nome da Obra',
            'item': 'C - Item/CPU - SIENGE',
            'unit': 'D - Unid serviço/Coef Insumo',
            'unitPrice': 'E - Preço Unit',
            'balance': 'J - Quant. Total JULHO 2026',
            'plannedQuantity': 'Qnt Realis.',
            'plannedValue': 'R$ Realista',
            'executedQuantity': 'Qtd. Executada',
            'executedValue': 'R$ Executado',
            'serviceOrder': 'Ordem de Serviço mensal replicada aos itens do contrato',
        },
    },
    'contractData': dict(contracts),
}

data_path.write_text(
    'window.SINAPE_MODEL_V2=' + json.dumps(model, ensure_ascii=False, separators=(',', ':')) + ';\n',
    encoding='utf-8',
)

if 'src="model-v2-data.js"' not in html:
    html = html.replace(
        '<script>!async function(){',
        '<script src="model-v2-data.js"></script><script>!async function(){',
        1,
    )

old_reader = 'function t(t){var n="var "+t+"="'
new_reader = 'function t(t){if(window.SINAPE_MODEL_V2&&Object.prototype.hasOwnProperty.call(window.SINAPE_MODEL_V2,t))return window.SINAPE_MODEL_V2[t];var n="var "+t+"="'
if old_reader in html:
    html = html.replace(old_reader, new_reader, 1)
elif new_reader not in html:
    raise RuntimeError('Função de leitura da base não localizada')

old_planned_value = 'function Z(e){return oe(!0).reduce(function(t,n){return t+Number(n.plan[e]||0)*n.price},0)}'
new_planned_value = 'function Z(e){return oe(!0).reduce(function(t,n){var a=n.planValue||[];return t+Number(a[e]||0)},0)}'
if old_planned_value in html:
    html = html.replace(old_planned_value, new_planned_value, 1)
elif new_planned_value not in html:
    raise RuntimeError('Cálculo do valor planejado não localizado')

html = html.replace('b=sessionStorage.getItem(f)||""', 'b=""', 1)

old_clear = 'document.getElementById("clearFilters").onclick=function(){c=d,s="Todas",m="",l=[],u="",Be.value=String(d),document.getElementById("osFilter").value="Todas",document.getElementById("statusFilter").value="",Ee(),we(),xe()}'
new_clear = 'document.getElementById("clearFilters").onclick=function(){L=C,q=T(L),c=-1,s="Todas",m="",l=[],u="",Ie.value=C,Be.value="-1",document.getElementById("itemFilter").value="",document.getElementById("osFilter").value="Todas",document.getElementById("statusFilter").value="",document.getElementById("advancedFilters").classList.remove("open"),ke(!1),Ee(),we(),xe()}'
if old_clear in html:
    html = html.replace(old_clear, new_clear, 1)

html = html.replace(
    'Fonte: Modelo Cronograma + Base BI Executado (2)',
    'Planejado e Executado: Modelo Cronograma Base de Dados v2 · RDO: Base BI Executado (2)',
)
html = html.replace(
    'Planejamento: base atual · Executado/RDO: <b>Base BI Executado (2) atualizada</b>',
    'Planejado, Executado e OS: <b>Modelo Cronograma Base de Dados v2</b> · RDO: Base BI Executado (2)',
)
html = html.replace('Fonte: Base BI Executado (2)', 'Fonte: Modelo Cronograma Base de Dados v2')
html = html.replace(
    'Planejamento: Modelo Cronograma Base de Dados · Planilha1',
    'Planejamento, Executado e OS: Modelo Cronograma Base de Dados v2 · base interna',
)
html = html.replace(
    'Executado e RDO: Estudo Geral · Base BI Executado (2)',
    'RDO: Estudo Geral · Base BI Executado (2) · somente quantidade',
)
html = html.replace(
    'Quantitativos e valores executados provenientes da Base BI Executado (2)',
    'Registros operacionais provenientes da Base BI Executado (2), sem exibição do valor financeiro',
)

html = html.replace(
    '<th>Contrato</th><th>Período</th><th>Observação do RDO</th><th>Serviço executado</th><th>Quantidade</th><th>Valor executado</th>',
    '<th>Contrato</th><th>Período</th><th>Observação do RDO</th><th>Serviço executado</th><th>Quantidade executada</th>',
)
html = html.replace(
    '+P.format(e.quantity)+" "+D(e.unit)+"</td><td><strong>"+A.format(e.value)+"</strong></td></tr>"',
    '+P.format(e.quantity)+" "+D(e.unit)+"</td></tr>"',
)

html_path.write_text(html, encoding='utf-8')

rows_2465 = contracts.get('2465 - DER SP', [])
sample = next((row for row in rows_2465 if row['name'].startswith('Oper. e Manutenção de Canteiro Tipo I')), None)
assert sample
assert abs(sample['price'] - 181544.96) < 0.01
assert abs(sample['balance'] - 0.9310882977412528) < 1e-9
assert abs(sample['exec'][0] - 0.08) < 1e-9
assert abs(sample['execValue'][0] - 14523.5968) < 0.01
assert all(abs(value) < 1e-12 for value in sample['plan'][6:])
print(json.dumps(model['metadata'], ensure_ascii=False, indent=2))
print('VALIDACAO_2465', json.dumps(sample, ensure_ascii=False))
