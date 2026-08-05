import json,sys,re,unicodedata
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import load_workbook

def norm(s):
    s='' if s is None else str(s)
    s=unicodedata.normalize('NFKD',s).encode('ascii','ignore').decode().lower()
    return re.sub(r'[^a-z0-9]+',' ',s).strip()

def parse_date(v):
    if hasattr(v,'strftime'): return v.strftime('%Y-%m-%d')
    s=str(v).strip()
    for f in ('%d/%m/%Y','%Y-%m-%d','%Y-%m-%d %H:%M:%S'):
        try:return datetime.strptime(s,f).strftime('%Y-%m-%d')
        except:pass
    raise ValueError(f'Data invalida: {v}')

xlsx=Path(sys.argv[1]); mapping_path=Path(sys.argv[2]); out=Path(sys.argv[3])
maprows=json.loads(mapping_path.read_text(encoding='utf-8'))
by_contract={}
for x in maprows: by_contract.setdefault(str(x['contract']),[]).append(x)
wb=load_workbook(xlsx,read_only=True,data_only=True)
if 'Executado' not in wb.sheetnames: raise SystemExit('Aba Executado nao encontrada')
ws=wb['Executado']
rows=list(ws.iter_rows(values_only=True))
header_i=next(i for i,r in enumerate(rows) if r and any(str(v).strip()=='Date' for v in r if v is not None))
headers=[str(v).strip() if v is not None else '' for v in rows[header_i]]
idx={h:i for i,h in enumerate(headers)}
required=['Date','Cód obra','Tarefa','Soma de Qtd']
for h in required:
    if h not in idx: raise SystemExit(f'Coluna obrigatoria ausente: {h}')
reg=[]; unmatched=[]
for r in rows[header_i+1:]:
    if not r or r[idx['Date']] in (None,''): continue
    raw_code=r[idx['Cód obra']]
    task=str(r[idx['Tarefa']] or '').strip()
    if raw_code in (None,'') or not task:
        continue
    code=str(raw_code).strip().split('.')[0]
    contract=code if len(code)>=4 else ('2'+code if len(code)==3 else code)
    qty=float(r[idx['Soma de Qtd']] or 0)
    candidates=by_contract.get(contract,[])
    nt=norm(task)
    exact=[x for x in candidates if norm(x['name'])==nt]
    if not exact:
        scored=[]
        ts=set(nt.split())
        for x in candidates:
            xs=set(norm(x['name']).split()); score=len(ts&xs)/max(1,len(ts|xs)); scored.append((score,x))
        scored.sort(key=lambda z:z[0],reverse=True)
        exact=[scored[0][1]] if scored and scored[0][0]>=0.72 else []
    if not exact:
        unmatched.append({'contrato':contract,'tarefa':task}); continue
    item=exact[0]; price=float(item.get('price') or 0)
    date=parse_date(r[idx['Date']])
    reg.append({'contract':contract,'month':int(date[5:7])-1,'itemId':item['itemId'],'observation':'Aba Executado - Estudo Geral','service':task,'unit':item.get('unit',''),'quantity':qty,'value':qty*price,'date':date})
if unmatched:
    Path(str(out)+'.nao_mapeados.json').write_text(json.dumps(unmatched,ensure_ascii=False,indent=2),encoding='utf-8')
    raise SystemExit(f'{len(unmatched)} linhas sem mapeamento; arquivo nao atualizado')
payload={'fonte':'Estudo Geral.xlsx > Executado','geradoEm':datetime.now(timezone.utc).isoformat(),'quantidadeRegistros':len(reg),'registros':reg}
out.write_text(json.dumps(payload,ensure_ascii=False,indent=2),encoding='utf-8')
print(f'OK: {len(reg)} registros')
